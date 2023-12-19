from flask import Flask, request, jsonify, send_file, Blueprint, url_for
from flask_cors import CORS, cross_origin
from datetime import datetime, timedelta
import os
import json
import helper, env

app = Flask(__name__, template_folder="view", static_folder="lib")
app.config["JSON_SORT_KEYS"] = False
app.json.sort_keys = False
app.secret_key = env.app_secret_key
app.config['CORS_HEADERS'] = 'Content-Type'
cors = CORS(app, resources={r"/*": {"origins": "*"}})

route = ""
user_data = None


ruled_auth_token = []
@app.before_request
def middleware():
    requested_route = request.path 
    if requested_route in ruled_auth_token:
        token = request.headers.get('auth-token')
        global user_data
        user_data = helper.db_raw(f"""
            SELECT * FROM _user WHERE USER_TOKEN = '{token}'
            """)[1]
        if len(user_data) == 0:
            return helper.composeReply("ERROR", "User invalid")
        user_data = user_data[0]


route = "/posyandu_get"
ruled_auth_token.append(route)
@app.route(route, methods=['POST'])
def posyandu_get():
    data = helper.db_raw("""
        SELECT * FROM posyandu
    """)[1]
    return helper.composeReply("SUCCESS", "Data Posyandu", data)


route = "/auth_register"
ruled_auth_token.append(route)
@app.route(route, methods=['POST'])
def auth_register():
    email = request.form.get("email")
    if email == None:
        return helper.composeReply("ERROR", "Parameter incomplete (email)")
    password = request.form.get("password")
    if password == None:
        return helper.composeReply("ERROR", "Parameter incomplete (password)")
    nama = request.form.get("nama")
    if nama == None:
        return helper.composeReply("ERROR", "Parameter incomplete (nama)")
    hp = request.form.get("hp")
    if hp == None:
        return helper.composeReply("ERROR", "Parameter incomplete (hp)")
    
    cek_user = helper.db_raw(f"""
            SELECT * FROM _user WHERE USER_EMAIL = '{email}'
        """)[1]
    print("cek_user", cek_user)
    if len(cek_user) > 0:
        return helper.composeReply("ERROR", f"Maaf, kader dengan email {email} sudah terdaftar")

    register = helper.db_insert("_user", {
        "USER_EMAIL" : email,
        "USER_PASSWORD_HASH" : password,
        "USER_NAMA" : nama,
        "USER_HP" : hp,
        "USER_TYPE" : "USER_TYPE_KADER"
    })
    
    return helper.composeReply("SUCCESS", "Berhasil menambahkan kader", register)


route = "/auth_login"
@app.route(route, methods=['POST'])
def auth_login():
    email = request.form.get("email")
    if email == None:
        return helper.composeReply("ERROR", "Parameter incomplete (email)")
    password = request.form.get("password")
    if password == None:
        return helper.composeReply("ERROR", "Parameter incomplete (password)")
    
    cek_user = helper.db_raw(f"""
            SELECT * FROM _user WHERE USER_EMAIL = '{email}'
        """)[1]
    print("cek_user", cek_user)
    if len(cek_user) == 0:
        return helper.composeReply("ERROR", f"Maaf, pengguna dengan email {email} tidak terdaftar")
    cek_user = cek_user[0]
    
    token = helper.generate_token()
    helper.db_update("_user",
                        {
                            "USER_TOKEN" : token
                        },
                        f"USER_ID = '{cek_user['USER_ID']}'"
                    )
    cek_user["USER_TOKEN"] = token
    
    return helper.composeReply("SUCCESS", "Login berhasil", cek_user)


route = "/tes_save"
ruled_auth_token.append(route)
@app.route(route, methods=['POST'])
def tes_save():
    posyandu = request.form.get("posyandu")
    if posyandu == None:
        return helper.composeReply("ERROR", "Parameter incomplete (posyandu)")
    tanggal = request.form.get("tanggal")
    if tanggal == None:
        return helper.composeReply("ERROR", "Parameter incomplete (tanggal)")
    nama = request.form.get("nama")
    if nama == None:
        return helper.composeReply("ERROR", "Parameter incomplete (nama)")
    jk = request.form.get("jk")
    if jk == None:
        return helper.composeReply("ERROR", "Parameter incomplete (jk)")
    umur = request.form.get("umur")
    if umur == None:
        return helper.composeReply("ERROR", "Parameter incomplete (umur)")
    tinggi = request.form.get("tinggi")
    if tinggi == None:
        return helper.composeReply("ERROR", "Parameter incomplete (tinggi)")
    berat = request.form.get("berat")
    if berat == None:
        return helper.composeReply("ERROR", "Parameter incomplete (berat)")
    kepala = request.form.get("kepala")
    if kepala == None:
        return helper.composeReply("ERROR", "Parameter incomplete (kepala)")
    hasil_tinggi = request.form.get("hasil_tinggi")
    if hasil_tinggi == None:
        return helper.composeReply("ERROR", "Parameter incomplete (hasil_tinggi)")
    hasil_berat = request.form.get("hasil_berat")
    if hasil_berat == None:
        return helper.composeReply("ERROR", "Parameter incomplete (hasil_berat)")
    hasil_kepala = request.form.get("hasil_kepala")
    if hasil_kepala == None:
        return helper.composeReply("ERROR", "Parameter incomplete (hasil_kepala)")

    add = helper.db_insert("tes", {
        "TES_POSY" : posyandu,
        "TES_TANGGAL" : tanggal,
        "TES_NAMA" : nama,
        "TES_JK" : jk,
        "TES_UMUR" : umur,
        "TES_TINGGI" : tinggi,
        "TES_BERAT" : berat,
        "TES_KEPALA" : kepala,
        "TES_HASIL_TINGGI" : hasil_tinggi,
        "TES_HASIL_BERAT" : hasil_berat,
        "TES_HASIL_KEPALA" : hasil_kepala
    })
    if not add[0]:
        return helper.composeReply("ERROR", "Internal server error occured, sorry")
    
    return helper.composeReply("SUCCESS", "Berhasil menyimpan pengecekkan", add)


route = "/tes_get"
ruled_auth_token.append(route)
@app.route(route, methods=['POST'])
def tes_get():
    posyandu = request.form.get("posyandu")
    if posyandu == None:
        return helper.composeReply("ERROR", "Parameter incomplete (posyandu)")
    qry = """
        SELECT * FROM tes
    """
    if posyandu != "_ALL_":
        qry += f" WHERE TES_POSY = {posyandu}"
    data = helper.db_raw(qry)[1]

    return helper.composeReply("SUCCESS", "Data pengecekkan", data)


route = "/file"
@app.route(route, methods = ["GET"])
def file():
    filename = request.args.get('filename')
    if not filename:
        return "N", 400
    
    image_path = f'images/{filename}'
    if image_path.endswith('.jpg') or image_path.endswith('.jpeg'):
        mimetype = 'image/jpeg'
    elif image_path.endswith('.png'):
        mimetype = 'image/png'
    else:
        return helper.composeReply("ERROR", 'Unsupported file type')
    return send_file(image_path, mimetype=mimetype)


route = "/export_simaset"
@app.route(route, methods=['GET'])
def export_simaset():
    opd = request.args.get("opd")
    if opd == None:
        return helper.composeReply("ERROR", "Parameter incomplete (opd)")
    export = request.args.get("export")
    if export == None:
        return helper.composeReply("ERROR", "Parameter incomplete (export)")
    qry = """
        SELECT A.*, B.*, C.R_VALUE AS BARANG_KONDISI_VALUE, D.R_VALUE AS BARANG_KEBERADAAN_VALUE FROM barang AS A
        JOIN dinas as B ON A.BARANG_OPD = B.DINAS_ID
        JOIN _reference as C ON A.BARANG_KONDISI = C.R_ID
        JOIN _reference as D ON A.BARANG_KEBERADAAN = D.R_ID
    """
    if not opd == "_ALL_":
        qry += f" WHERE A.BARANG_OPD = {opd}"
    data = helper.db_raw(qry, database="simaset")[1]

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from io import BytesIO

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'

    row_start = 1
    row_current = row_start

    row_current += 1
    display_judul = "DATA BARANG"
    if not opd == "_ALL_":
        display_judul += "" # tambah ket. OPD
    sheet[f"B{row_current}"] = display_judul
    sheet[f"B{row_current}"].font = Font(bold=True, size=14)
    sheet[f"B{row_current}"].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(f"B{row_current}:H{row_current}")

    row_current += 2
    sheet[f"B{row_current}"] = "Kode Sensus"
    sheet[f"C{row_current}"] = "Nama Barang"
    sheet[f"D{row_current}"] = "OPD"
    sheet[f"E{row_current}"] = "Kondisi"
    sheet[f"F{row_current}"] = "Keberadaan"
    sheet[f"G{row_current}"] = "Waktu Pendataan"
    sheet[f"H{row_current}"] = "Keterangan"
    for cell in helper.get_cells_in_range(sheet, f"B-{row_current}:H-{row_current}"):
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="c4c1c0", end_color="c4c1c0", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sheet.row_dimensions[row_current].height = 20
    sheet.column_dimensions[("B")].width = 20
    sheet.column_dimensions[("C")].width = 20
    sheet.column_dimensions[("D")].width = 20
    sheet.column_dimensions[("E")].width = 20
    sheet.column_dimensions[("F")].width = 20
    sheet.column_dimensions[("G")].width = 20
    sheet.column_dimensions[("H")].width = 20
    
    row_data_start = row_current + 1
    for i, record in enumerate(data):
        row_current += 1
        sheet[f"B{row_current}"] = record["BARANG_KODE_SENSUS"]
        sheet[f"C{row_current}"] = record["BARANG_NAMA"]
        sheet[f"D{row_current}"] = record["DINAS_NAMA"]
        sheet[f"E{row_current}"] = record["BARANG_KONDISI_VALUE"]
        sheet[f"F{row_current}"] = record["BARANG_KEBERADAAN_VALUE"]
        sheet[f"G{row_current}"] = helper.tgl_indo(record["BARANG_WAKTU_PENDATAAN"], 'SHORT')
        sheet[f"H{row_current}"] = record["BARANG_KETERANGAN"]
    row_data_end = row_current
    for cell in helper.get_cells_in_range(sheet, f"B-{row_data_start}:H-{row_data_end}"):
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_current += 2
    current_datetime = helper.get_local_time().strftime('%Y-%m-%d %H:%M:%S')
    sheet[f"B{row_current}"] = f"diunduh pada : {helper.tgl_indo(current_datetime, 'LONG')}"
    sheet[f"B{row_current}"].font = Font(italic=True)
    sheet.merge_cells(f"B{row_current}:C{row_current}")

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    if export == "pdf":
        print(1)
    else:
        return send_file(excel_file, as_attachment=True, download_name=f'data barang {current_datetime}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    

route = "/export_kmsbalita"
@app.route(route, methods=['GET'])
def export_kmsbalita():
    posyandu = request.args.get("posyandu")
    if posyandu == None:
        return helper.composeReply("ERROR", "Parameter incomplete (posyandu)")
    posyandu_nama = request.args.get("posyandu_nama")
    if posyandu_nama == None:
        return helper.composeReply("ERROR", "Parameter incomplete (posyandu_nama)")
    export = request.args.get("export")
    if export == None:
        return helper.composeReply("ERROR", "Parameter incomplete (export)")
    qry = """
        SELECT A.*, B.*, C.R_VALUE AS TES_JK_VALUE FROM tes AS A
        JOIN posyandu as B ON A.TES_POSY = B.POSY_ID
        JOIN _reference as C ON A.TES_JK = C.R_ID
    """
    if not posyandu == "_ALL_":
        qry += f" WHERE A.TES_POSY = {posyandu}"
    data = helper.db_raw(qry)[1]
    print(data)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from io import BytesIO

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'

    row_start = 1
    row_current = row_start

    row_current += 1
    display_judul = "DATA PEMERIKSAAN"
    if not posyandu == "_ALL_":
        display_judul += "" # tambah ket. OPD
    sheet[f"B{row_current}"] = display_judul
    sheet[f"B{row_current}"].font = Font(bold=True, size=14)
    sheet[f"B{row_current}"].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(f"B{row_current}:H{row_current}")

    row_current += 1
    sheet[f"B{row_current}"] = f"Posyandu : {posyandu_nama}"
    sheet[f"B{row_current}"].font = Font(bold=True)

    row_current += 2
    sheet[f"B{row_current}"] = "Nama"
    sheet[f"C{row_current}"] = "Jenis"
    sheet[f"D{row_current}"] = "Umur"
    sheet[f"E{row_current}"] = "Tinggi"
    sheet[f"F{row_current}"] = "Berat"
    sheet[f"G{row_current}"] = "Lingkar Kepala"
    sheet[f"H{row_current}"] = "Tanggal Pengecekkan"
    for cell in helper.get_cells_in_range(sheet, f"B-{row_current}:H-{row_current}"):
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="c4c1c0", end_color="c4c1c0", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sheet.row_dimensions[row_current].height = 20
    sheet.column_dimensions[("B")].width = 20
    sheet.column_dimensions[("C")].width = 20
    sheet.column_dimensions[("D")].width = 20
    sheet.column_dimensions[("E")].width = 20
    sheet.column_dimensions[("F")].width = 20
    sheet.column_dimensions[("G")].width = 20
    sheet.column_dimensions[("H")].width = 25
    
    row_data_start = row_current + 1
    for i, record in enumerate(data):
        row_current += 1
        sheet[f"B{row_current}"] = record["TES_NAMA"]
        sheet[f"C{row_current}"] = record["TES_JK_VALUE"]
        sheet[f"D{row_current}"] = f"{record['TES_UMUR']} bulan"
        sheet[f"E{row_current}"] = f"{record['TES_TINGGI']} cm ({record['TES_HASIL_TINGGI']})"
        sheet[f"F{row_current}"] = f"{record['TES_BERAT']} kg ({record['TES_HASIL_BERAT']})"
        sheet[f"G{row_current}"] = f"{record['TES_KEPALA']} cm ({record['TES_HASIL_KEPALA']})"
        sheet[f"H{row_current}"] = helper.tgl_indo(record["TES_TANGGAL"], 'SHORT').replace("00:00:00", "")
    row_data_end = row_current
    for cell in helper.get_cells_in_range(sheet, f"B-{row_data_start}:H-{row_data_end}"):
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_current += 2
    current_datetime = helper.get_local_time().strftime('%Y-%m-%d %H:%M:%S')
    sheet[f"B{row_current}"] = f"diunduh pada : {helper.tgl_indo(current_datetime, 'LONG')}"
    sheet[f"B{row_current}"].font = Font(italic=True)
    sheet.merge_cells(f"B{row_current}:C{row_current}")

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    if export == "pdf":
        from jinja2 import Environment, FileSystemLoader
        import pdfkit

        env = Environment(loader=FileSystemLoader('view'))
        template = env.get_template('template.html')

        row = ""
        for i, record in enumerate(data):
            row += f"""
                <tr>
                    <td>{record["TES_NAMA"]}</td>
                    <td>{record["TES_JK_VALUE"]}</td>
                    <td>{record['TES_UMUR']} bulan</td>
                    <td>{record["TES_TINGGI"]} cm<br>({record['TES_HASIL_TINGGI']})</td>
                    <td>{record["TES_BERAT"]} kg<br>({record['TES_HASIL_BERAT']})</td>
                    <td>{record["TES_KEPALA"]} cm<br>({record['TES_HASIL_KEPALA']})</td>
                    <td>{helper.tgl_indo(record["TES_TANGGAL"], 'SHORT').replace("00:00:00", "")}</td>
                </tr>
            """
        rendered_html = template.render(row=row,
                                        title=display_judul,
                                        subtitle=f"Posyandu : {posyandu_nama}",
                                        downloaded=f"diunduh pada : {helper.tgl_indo(current_datetime, 'LONG')}",
                                        img1=url_for('file', _external=True) + "?filename=img.png"
                                        )

        with open('temp.html', 'w', encoding='utf-8') as html_file:
            html_file.write(rendered_html)

        filename = f'storage/DATA PEMERIKSAAN POSYANDU {posyandu_nama}.pdf'
        pdfkit.from_file('temp.html', filename)
        response = send_file(filename, as_attachment=True)
        
        os.remove("temp.html")
        #os.remove(filename)
        return response
    else:
        return send_file(excel_file, as_attachment=True, download_name=f'DATA PEMERIKSAAN POSYANDU {posyandu_nama}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
   


if __name__ == '__main__':
    app.run(host = env.runHost, port = env.runPort, debug = env.runDebug)